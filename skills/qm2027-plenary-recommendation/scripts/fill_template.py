#!/usr/bin/env python3
"""Fill the QM2027 IAC plenary/lecturer recommendation xlsx from extracted email data.

Usage:
    python3 fill_template.py --template <template.xlsx> --data <data.json> --out <output.xlsx>

data.json shape:
[
  {
    "sheet": "plenary" | "lecturer",
    "recommender": "Laura Tolos (ICE, CSIC)",
    "candidates": [
      {"priority": 1, "name": "...", "affiliation": "...", "gender": "",
       "topic": "...", "nationality": "...", "remarks": ""},
      ...
    ]
  },
  ...
]

Rules encoded here (see SKILL.md for the full explanation):
  - Each recommender occupies one 3-row block (priority rows 1/2/3) in the target sheet.
  - If a recommender's email did not state any ranking, every candidate's priority is
    forced to 1 (per user instruction: "순위가 안 적혀 있으면 모두 1순위").
  - If a recommender nominated fewer than 3 people, the unused rows in that block are
    left blank.
  - Unknown fields are left blank ("" or omitted) rather than guessed.
  - Sheet "plenary" -> workbook sheet "IAC Recommendations" (blocks start at row 3).
  - Sheet "lecturer" -> workbook sheet "Lecturuer" (blocks start at row 9).
  - Recommenders beyond the number of blank blocks available in the template raise an
    error so a human can extend the template rather than silently dropping data.
"""
import argparse
import json
import sys

import openpyxl

SHEET_CONFIG = {
    "plenary": {"sheet_name": "IAC Recommendations", "first_block_row": 3},
    "lecturer": {"sheet_name": "Lecturuer", "first_block_row": 9},
}

FIELDS = ["priority", "name", "affiliation", "gender", "topic", "nationality", "remarks"]
COLUMNS = {"priority": 2, "name": 3, "affiliation": 4, "gender": 5, "topic": 6, "nationality": 7, "remarks": 8}


def find_blank_blocks(ws, first_block_row, block_size=3):
    """Yield starting row numbers of 3-row blocks whose Name column (C) is empty."""
    row = first_block_row
    while True:
        name_cells = [ws.cell(row=row + i, column=3).value for i in range(block_size)]
        priority_cells = [ws.cell(row=row + i, column=2).value for i in range(block_size)]
        # A real block has Priority pre-filled (1,2,3) even when blank of data.
        if not any(p is not None for p in priority_cells):
            break
        if all(n is None for n in name_cells):
            yield row
        row += block_size


def fill_block(ws, start_row, recommender, candidates):
    ws.cell(row=start_row, column=1, value=recommender)
    ranked = any(c.get("priority") for c in candidates)
    for i, cand in enumerate(candidates[:3]):
        r = start_row + i
        priority = cand.get("priority")
        if not ranked or priority is None:
            priority = 1
        ws.cell(row=r, column=COLUMNS["priority"], value=priority)
        for field in ("name", "affiliation", "gender", "topic", "nationality", "remarks"):
            val = cand.get(field, "") or ""
            ws.cell(row=r, column=COLUMNS[field], value=val)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    with open(args.data, encoding="utf-8") as f:
        entries = json.load(f)

    wb = openpyxl.load_workbook(args.template)

    # available blank blocks per sheet, consumed in order
    block_iters = {
        key: list(find_blank_blocks(wb[cfg["sheet_name"]], cfg["first_block_row"]))
        for key, cfg in SHEET_CONFIG.items()
    }

    for entry in entries:
        sheet_key = entry["sheet"]
        cfg = SHEET_CONFIG[sheet_key]
        ws = wb[cfg["sheet_name"]]
        blocks = block_iters[sheet_key]
        if not blocks:
            print(
                f"ERROR: no more blank recommender blocks left in sheet "
                f"'{cfg['sheet_name']}' for recommender '{entry['recommender']}'. "
                f"Extend the template with more 3-row blocks and retry.",
                file=sys.stderr,
            )
            sys.exit(1)
        start_row = blocks.pop(0)
        fill_block(ws, start_row, entry["recommender"], entry["candidates"])

    wb.save(args.out)
    print(f"saved: {args.out}")


if __name__ == "__main__":
    main()
